import os
import subprocess
from itertools import dropwhile
from functools import reduce
import yaml
from collections import Counter
import pandas as pd
from sys import argv
import os.path
from os.path import join
import numpy as np
import pickle as pkl
from json import JSONEncoder, JSONDecoder
import re
import warnings
from openpyxl import load_workbook
import logging
from functools import lru_cache
import pprint

from src.pytypes import *
from src import helpers as h
from src.helpers import clean_str


logger = logging.getLogger("LEEM")

maquettefile = os.path.realpath(h.last_modified_file("data/maquette"))
cahierdescharges_file = os.path.realpath(h.last_modified_file("data/cahier_des_charges"))
# maquettefile = os.path.realpath("data/Maquette LEEM HSS pour DeSI")
# cahierdescharges_file = os.path.realpath("data/Cahier des Charges DeSI")
skilldescriptionfile = os.path.realpath("data/skill_description.xlsx")
jobselectionfile = os.path.realpath("data/Fiches métiers MD240920.xlsx")

pdf_dir = 'data/LEEM_pdfs'

def easyspaces(s):
    return re.sub('\s+', ' ', s).strip()

def load_file(f):
    print(f"loading {f}...")
    l = subprocess.check_output(["pdf2txt", f]).decode("utf-8").split("\n")
    l = [easyspaces(s) for s in l]
    return l


def is_empty(e):
    return e.strip() == ""


def span(predicate, l):
    def f(la,lb):
        if lb == []:
            return la, lb
        else:
            x, *xs = lb
            if predicate(x):
                return f(la+[x], xs)
            else:
                return la, lb
    return f([], l)


def find_start(l):
    return list(dropwhile(lambda x: "COMPÉTENCES CLÉS" not in x, l))[1:]


def read_title(l):
    a1, t, a2, *xs = l
    if t.lower() == "anglais":
        return t, [a2]+xs
    else:
        if not ( is_empty(a1) and is_empty(a2) ):
            return None, l
        else:
            return t, xs


def find_body(l):
    b, xs = span(lambda e: not is_empty(e), l)
    def ff(a, b):
        b = easyspaces(b)
        return (a+[b]) if b.startswith("- ") else (a[:-1] + [a[-1] + ' ' + b])
    return reduce(ff, b, []), xs

    
def find_content(l):
    def is_bs_line(k):
        return len(l[k])==0 and l[k-1].startswith("-") and l[k+1].startswith("-")
    lbs = [i for i in range(len(l)) if is_bs_line(i)]
    for k in reversed(lbs):
        del(l[k])
    
    def f(ll):
        try:
            t, lll = read_title(ll)
            if t is None:
                return [(None, ll)]
            else:
                b, llll = find_body(lll)
                return [(t, b)] + f(llll)
        except Exception as e:
            return []
    ll = f([ e.strip() for e in find_start(l)])
    return {k: v for k, v in ll if k is not None and len(v)>0}





def filter_keys(d, f):
    return { k: d[k] for k in d.keys() if f(k) }


def merge_anglais(ts):
    l = [i for i, e in enumerate(ts.keys()) if e.lower().startswith("anglais")]
    def getf(i):
        return ts[ list(ts.keys())[i] ]
        
    for e in l:
        getf(e-1).extend( [ list(ts.keys())[e] + " : " + ee for ee in getf(e) ] )
    
    for i in reversed(l):
        del ts[list(ts.keys())[i]]
    
    return ts

    
    
def walk_all(start_dir = pdf_dir, only=None):
    def f(start_dir):
        filename = start_dir.split("/")[-1]
        print(start_dir)
        if os.path.isfile(start_dir):
            if start_dir.endswith(".pdf") and (only is None or filename in only):
                o = find_content(load_file(start_dir))
                o = merge_anglais(o)
                return Leaf( start_dir.replace(pdf_dir+'/', ''), o )
        else:
            print("not a file")
            p, ds, fs = next(os.walk(start_dir))
            n = Node( filename, list(filter(lambda e: not e is None, [ f(join(start_dir, e)) for e in ds+fs ])))
            if len(n.children):
                return n
    # d = f(start_dir)
    # with open(dest, "w", encoding="utf-8") as f:
    #     yaml.dump(d, f, encoding="utf-8", allow_unicode=True)
    return f(start_dir)



def flatten_skills(d, prefix=""):
    # print(f"a: {prefix=}")
    def f(k, prefix):
        # print(f"b: {prefix=}")
        prefix = os.path.join(prefix, k)
        if k.endswith(".pdf"):
            return [(prefix, d[k])]
        else:
            return flatten_skills(d[k], prefix)
    
    return [ e for ee in [f(ab, prefix) for ab in d.keys()] for e in ee]


def combine_skills(acc, b):
    def concat_dict_keys(k):
        return acc.get(k, []) + b.get(k, [])
    
    return {k: concat_dict_keys(k) for k in set(list(acc.keys()) + list(b.keys()))}



def each_skill_to_excel(ddd, filepath = "results/", diffWith=maquettefile):
    # logging.warning(f"saving to {filepath}")
    with pd.ExcelWriter(filepath) as ex:
        
        def to_df(key):
            sheet_name = f"Compétences {key}"
            
            tdf = pd.read_excel(diffWith, sheet_name=sheet_name)
            tdf = tdf[[sheet_name, "Soft-Skills (Combinaison)"]].dropna()
            
            items, counts = np.unique(ddd[key], return_counts=True)
            df = pd.DataFrame(zip(counts, items), columns=["n", sheet_name])
            for e in "Hard-Skills,Soft-Skills (Combinaison),PERSONNALITE (composante brute)".split(','):
                df[e] = ""
            
            ln = df.n
            df = df.merge(tdf, on=sheet_name, how="left")
            df['n'] = ln
            df.drop(columns=["Soft-Skills (Combinaison)_x"])
            df = df[["n",sheet_name,"Hard-Skills","Soft-Skills (Combinaison)_y","PERSONNALITE (composante brute)"]]
            df = df.rename(columns={"Soft-Skills (Combinaison)_y": "Soft-Skills (Combinaison)"})
            
            df = df.sort_values("n", ascending=False)
            
            df.to_excel(ex, sheet_name=sheet_name, index=False)
            print(ex.sheets[sheet_name].__dict__.keys())
            # for e in ex.sheets[sheet_name].column_dimensions["A"]:
            #     print(e.width)
            ex.sheets[sheet_name].column_dimensions["A"].width = 2
            ex.sheets[sheet_name].column_dimensions["B"].width = 77
            ex.sheets[sheet_name].column_dimensions["C"].width = 8
            ex.sheets[sheet_name].column_dimensions["D"].width = 47
            ex.sheets[sheet_name].column_dimensions["E"].width = 10
            # ex.sheets[sheet_name].set_column(0, 10, 100)
            
        for k in sorted(ddd.keys()):
            to_df(k)


@lru_cache()
def load_jobselectionnames(path=jobselectionfile, which=None):
    df = pd.read_excel(path).dropna()

    if which is not None:
        df = df.loc[[which in clean_str(e) for e in df.included] , :]
        
    return df.job_path


def get_desired_jobs_for_mockup(d, which=None, reference_file=jobselectionfile):
    l = load_jobselectionnames(path=reference_file, which=which)
    return [e for e in d if any([e.name.endswith(path) for path in l])]


def extraction_desired_skillf_for_mockup(d, reference_file=jobselectionfile):
    tout = get_desired_jobs_for_mockup(d, "tout")
    transverse = get_desired_jobs_for_mockup(d, "transvers")
    
    tout_skills = [e.content for e in tout]
    transverse_skills = [dict(Transverses=e.content["Transverses"]) for e in transverse]
    
    ans = reduce(combine_skills, tout_skills + transverse_skills)
    each_skill_to_excel(ans, "results/competences_to_sskill.xlsx")

    ## create job_status_usage (not needed after a first creation)
    # df = pd.DataFrame()
    # df["job_path"] = [e.name for e in d.find_all()]

    # all_skills_included = [e.name for e in (med+extra)]
    # only_transverse_included = [e.name for e in others]
    # def f(e):
    #     if e in all_skills_included:
    #         return "tout"
    #     if e in only_transverse_included:
    #         return "transverses"
    #     else:
    #         return ""
    # df["included"] = [ f(e) for e in df.job_path ]
    # df.to_excel("results/job_status_usage_for_sskills.xlsx", index=False)

    
def load_data(start_dir="data/LEEM_pdfs", cache="results/skills.yaml", only=None):
    '''
    loads the data of the PDFs, scraped and extracted in a Tree (cf Types).
    the scraping is cached so it is required to delete the cache whenever the PDF content gets updated
    '''
    print("******************* loading data from:", os.getcwd(), "***********************")
    print( os.listdir() )
    if not os.path.exists(cache) or not cache:
        d = walk_all(start_dir=start_dir, only=None)
        # with open("data/jobs.pkl", "wb") as f:
        #     pkl.dump(d, f)
        with open(cache, "w", encoding="utf-8") as f:
            yaml.dump(d, f, encoding="utf-8", allow_unicode=True)
    else:
        with open(cache, "r", encoding="utf-8") as f:
            d = yaml.load(f)
        print("loading from cached version")
    if only is not None:
        d = d.filter(lambda e: e.name in only)
        # with open("data/jobs.pkl", "rb") as f:
        #     d = pkl.load(f)
    return d


@lru_cache()
def load_sskills_aggregator(filename=maquettefile):
    '''
    loads the maquette file, reads all the tabs and mix'em all together,
then returns a dict of { skill: [softskill] } 
'''
    axel = pd.ExcelFile(filename)
    def splitf(e):
        # print("splitf", e)
        return [clean_str(ee) for ee in h.splitter(e, ", ;".split())]
    def f(t):
        df = axel.parse(t).iloc[:, [1,3]].dropna().reindex()
        df.iloc[:, 1] = [ splitf(e) for e in df.iloc[:, 1] ]
        return zip(*[list(df[e]) for e in df.columns])
    # dfs = [ ee for t in axel.sheet_names for ee in f(t) ]
    dfs = [ ee for t in axel.sheet_names if t.startswith('Compétences ') for ee in f(t)]
    dic = { k: v for k,v in dfs }
    return dic


@lru_cache()
def load_sskill_description(filename=skilldescriptionfile):
    df = pd.read_excel(filename, sheet_name="sskill_description")

    for c in df.columns:
        df[c] = [e.strip() for e in df[c]]
    
    return df


@lru_cache()
def load_tp(filename=cahierdescharges_file):
    df = pd.read_excel(filename, sheet_name="Questions Personnalité", skiprows=[0,1,2,3], header=None, usecols='C,K,L').dropna()
    df.columns = "description name reversed".split()
    for c in df.columns[:-1]:
        df[c] = [e.strip() for e in df[c]]
    df["reversed"] = df.reversed == 0
    # df
    # d = {e['name']: e for e in df.to_dict("records")}
    
    return df


@lru_cache()
def load_tp_ss_corresp(filename=cahierdescharges_file):
    ws = [e for e in load_workbook(filename, data_only=True).worksheets if e.title=="Référentiel SS"][0]
    c = ws.cell(15,9)
    dftp = load_tp()
    tps = [clean_str(e) for e in dftp.name]
    d = {}
    for r in range(3,267):
        def readcell(ri,ci):
            cell = ws.cell(ri,ci)
            value = cell.value
            
            if value is not None :
                value = value.replace('\xa0',' ').strip()
                return dict(
                    value=value,
                    double_underline = cell.font.u == 'double'
                )
            else:
                return None
        l = [e for e in [readcell(r, c) for c in [7,9,11,13,15,17]] if e is not None and clean_str(e['value']) in tps]
        if len(l):
            d[readcell(r, 3)["value"].replace("- ", "").strip()] = l
    
    # df = pd.read_excel(filename, sheet_name="Référentiel SS", skiprows=[0,1], header=None, usecols='C,G,I,K,M').dropna()
    return d


def add_sskills(o):
    aggregadict = load_sskills_aggregator()
    def aggregate_l(l):
        try:
            l = [e for e in l if e in aggregadict.keys()]
            return list(sorted(
                zip(*np.unique( [ v for e in l for v in aggregadict[e]], return_counts=True)),
                key = lambda e: e[1], reverse=True
            ))
        except Exception as e:
            raise(e)
    def normalize_sskills_job(l):
        if len(l)==0:
            return []
        maxs = sorted(list([e[::-1] for e in zip(* np.unique([e[1] for e in l], return_counts=True))]), reverse=True)
        maxs_uniques = list(sorted(np.unique([e[1] for e in maxs]), reverse = True))
        def ff(e):
            k = [ ee[1] for ee in maxs if ee[1]==e[1] ][0]
            
            if k == maxs_uniques[0]:
                ans = 70
            elif k == maxs_uniques[1]:
                ans = 60
            else:
                ans = 50
            return ans
        return list(zip( [e[0] for e in l], [ff(e) for e in l] ))
    # return dict( skills=o, aggregated=normalize_sskills_job(aggregate_l([ee for e in o.values() for ee in e]) ) )
    ## o has Métier and Transverse, add _all key for the concatenation of all competences
    o['_all'] = [ee for e in o.values() for ee in e]
    
    return dict( skills=o, aggregated={
        k: normalize_sskills_job(aggregate_l(v)) for k, v in o.items()} )



def sskill_score_list(job, results, which="_all"):
    # print('\n\n===========================this is the job:\n\n', job)
    # print(type(job))
    # print(job.content)
    dr = {clean_str(k): v for k, v in results}
    dj = [ (k,v) for k, v in job.content["aggregated"][which] if dr.get(k, None) is not None]
    #l ogging.warning(dr)
    
    return { k: (dr.get(clean_str(k), None), v) for k, v in dj }


def sskill_score(job, results, which="_all"):
    debug_job = "Dev clinique/Redacteur medical.pdf"
    def llog(e):
        if job.name==debug_job:
            pprint.pprint(e)
    
    if len(job.content['aggregated'][which]) == 0:
        return None
    sq = lambda e: e*e
    results = [ (clean_str(a), b) for a,b in results ]
    # l = [ sq(dr.get(k, 9999)-v) for k, v in dj if dr.get(k) is not None]
    l = sskill_score_list(job, results, which).items()
    llog("*********************************++++++++")
    llog(('results', results))
    # llog(('job', job, job.content["aggregated"]))
    llog(('l', l))
    
    #print(results)
    llog('##############')

    if len(l)==0:
        logging.debug(f"{job} is not possible because there is no correspondance of sskills")
        return "no correspondance between skills and job"
    
    for e, v in job.content['aggregated'][which]:
        if v >= 60 and e not in [e[0] for e in results]:
            logging.debug(f"{job} is not possible because '{e}' is not in the results")
            # logging.debug(str(results))
            return f"'{e}' is not present in the results of the questionnaire"
        
    for ssk, (score, job_score) in l:
        if job_score >= 60 and score == 0:
            logging.debug(f"{job} is not possible because no competence for '{ssk}' ")
            return f"no competence for '{ssk}'"

    l = [e[1] for e in l]
    
    if sum([ e[0] for e in l ]) == 0:
        logging.debug('the sum of everything is zero')
        return "empty list"

    # if job.name.startswith("Responsable conseil"):
    #     print(l)
    #     print(len(job.content["aggregated"]))
    if len(l)>0:
        l = [ max(0, 100*min(a, b)/b) for a, b in l ]
        ans = np.mean(l)
        # ans = 1/(1+np.mean(l))
        # print(f"========>>>>>>> JOB {job.content} {ans}")
        # logging.debug(f"{job} is ok with {ans}")
        return ans


def sskills_from_tp(tp):
    ## reverse what should be reversed in the tp questionnaire
    for k in tp.keys():
        tp[k] *= 10
        try:
            reversed = personality_traits.loc[personality_traits.name == k].iloc[0]['reversed']
            if reversed:
                tp[k] = 100-tp[k]
        except Exception as ex:
        # if k not in personality_traits.keys():
            logging.error(f"{personality_traits}")
            logging.error(k)
            logging.error(f"{personality_traits.keys()}")
            raise ex
        
    
    
    def f(e):
        try:
            v = tp[e["value"]]
            return 100-v if e["double_underline"] else v
        except Exception as ex:
            #logging.warning("---- exception ----")
            #print(ex)
            #traceback.print_exc()
            #logging.warning('$$$$$$$$')
            #print(f'e: {e}')
            #print(f"tp {tp}")
            return None
    
    ss_scores = {}
    for ss in tp_ss_corresp.keys():
        l = [ e for e in [f(e) for e in tp_ss_corresp[ss]] if e is not None]
        ss_scores[ss] = np.mean(l) if len(l)>0 else None
        # print(f"ss: {ss} -> {l} -> {ss_scores[ss]}")
        
    # logging.debug(f"SS_SCORES ====>>>> {ss_scores}")
    return ss_scores


@lru_cache()
def load_aggregated_joblist(start_dir='data/LEEM_pdfs', cache="results/skills.yaml", only=None):
    data = load_data(start_dir, cache=cache, only=only)
    data = data.map(add_sskills).flatten()
    return data


def show_description(l, title):
    def f(g):
        return round(g(l), 2)
    
    hist=list(zip(*np.unique(l, return_counts=True)))
    plot = "\n".join([f"{k:3} {'#'*n}" for k, n in hist])
    logger.info(f'''

Répartition du nombre de {title}: {f(np.mean)}+/-{f(np.std)} (median={f(np.median)}, minmax={min(l)} - {max(l)})

Diagramme en barres:
{plot}


''')


def __init__():
    global personality_traits, tp_ss_corresp
    personality_traits = load_tp()
    tp_ss_corresp = load_tp_ss_corresp()

    
def describe_data(personality_traits, tp_ss_corresp, jobs):
    # jobs = load_aggregated_joblist()
    pd.set_option('display.max_rows', 500)
    pd.set_option('display.max_columns', 500)
    pd.set_option('display.width', 1000)
    logger.info(f"""


------------------------------------------------
---------------- extractor init ----------------
------------------------------------------------

{maquettefile=}
{cahierdescharges_file=}
{skilldescriptionfile=}
{jobselectionfile=}

""")
    
    ############
    logger.info(f'''
{len(personality_traits)} traits de personnalité.
{len(tp_ss_corresp)} soft skills.
{len(jobs)} jobs ({len([ e for e in jobs if len(e.content['aggregated'])==0 ])} jobs have no detected skills)

''')
    
    show_description([len(e.content['aggregated']) for e in jobs], "compétences pour les différents jobs")
    def unlines(l):
        return "\n".join(l)
    onlymockup = get_desired_jobs_for_mockup(jobs)
    dfjobselect = load_jobselectionnames()
    logger.info(f"""
-----------------------------------
{len(onlymockup)} jobs sélectionnés manuellement:
{dfjobselect.reset_index().drop(columns="index")}
""")
    show_description([len(e.content['aggregated']) for e in onlymockup], "compétences pour les différents jobs manuellement sélectionnés")





    

if __name__ == "__main__":
    d = load_data()
    extraction_desired_skillf_for_mockup(d)
    # walk_all(argv[1], dest = argv[2])
    
    # dd = flatten_skills(d)
    
    # ## fix the weird stuff
    # for i, e in enumerate(dd):
    #     if len(e[1].keys())>2:
    #         print(i, e[0], e[1].keys())
    #         merge_anglais(dd[i][1])

    # ddd = reduce(combine_skills, map(lambda e: e[1], dd))[1]
    
    # len(set(ddd['Transverses']))
    # len(ddd['Transverses'])
    # len(ddd['Métier'])
    # len(set(ddd['Métier']))

